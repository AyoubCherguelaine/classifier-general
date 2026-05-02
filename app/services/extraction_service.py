from pathlib import Path

import docx2txt
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader
import pytesseract

from app.core.exceptions import ExtractionError


DOC_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff"}
TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json"}


class ExtractionService:
    @staticmethod
    def _extract_pdf(file_path: Path, first_page_only: bool = False) -> str:
        reader = PdfReader(str(file_path))
        chunks: list[str] = []
        pages = reader.pages[:1] if first_page_only else reader.pages
        for page in pages:
            text = page.extract_text() or ""
            if text.strip():
                chunks.append(text)
        return "\n".join(chunks)

    @staticmethod
    def _extract_docx(file_path: Path) -> str:
        return docx2txt.process(str(file_path))

    @staticmethod
    def _extract_xlsx(file_path: Path) -> str:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if row_values:
                    chunks.append(" ".join(row_values))
        workbook.close()
        return "\n".join(chunks)

    def extract_text(self, file_name: str, file_path: Path, pdf_first_page_only: bool = False) -> str:
        extension = Path(file_name).suffix.lower()

        try:
            if extension in DOC_EXTENSIONS:
                if extension == ".pdf":
                    return self._extract_pdf(file_path, first_page_only=pdf_first_page_only)
                if extension == ".docx":
                    return self._extract_docx(file_path)
                if extension == ".xlsx":
                    return self._extract_xlsx(file_path)

            if extension in IMAGE_EXTENSIONS:
                image = Image.open(file_path)
                return pytesseract.image_to_string(image)

            if extension in TEXT_EXTENSIONS:
                return file_path.read_text(encoding="utf-8", errors="ignore")

            raise ExtractionError(f"Unsupported file extension: {extension or 'unknown'}")
        except ExtractionError:
            raise
        except Exception as exc:
            raise ExtractionError("Failed to extract text from file") from exc


extraction_service = ExtractionService()
